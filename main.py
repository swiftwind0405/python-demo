from __future__ import annotations

import argparse
import logging
import sys
from collections import OrderedDict
from dataclasses import dataclass
from datetime import date as DateType
from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

try:
    import tkinter as tk
    from tkinter import filedialog
except ImportError:  # pragma: no cover - Tkinter might be unavailable
    tk = None
    filedialog = None

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils.datetime import CALENDAR_MAC_1904, CALENDAR_WINDOWS_1900, from_excel


DAY_MINUTES = 24 * 60
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_ROOT = BASE_DIR / "outputs"
REST_BREAKS: Tuple[Tuple[int, int, str], ...] = (
    (11 * 60 + 30, 12 * 60, "11:30-12:00"),
    (17 * 60, 17 * 60 + 30, "17:00-17:30"),
    (23 * 60, 23 * 60 + 30, "23:00-23:30"),
)
REST_LABELS = {"一线员工休息"}
TARGET_RULES = {
    1: "full",
    2: "full",
    3: "until_midnight",
    5: "after_midnight",
    6: "until_midnight",
}
ERROR_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
LOGGER = logging.getLogger("overtime_calculator")
LOG_HANDLER_INITIALIZED = False


@dataclass(frozen=True)
class OvertimeResult:
    hours: float
    rest_labels: Tuple[str, ...]
    shift_text: str
    error: Optional[str] = None


def parse_time_to_minutes(token: str) -> Optional[int]:
    token = token.strip()
    if not token:
        return None
    if ":" not in token:
        return None
    hour_part, minute_part = token.split(":", 1)
    try:
        hour = int(hour_part)
        minute = int(minute_part)
    except ValueError:
        return None
    if hour == 24 and minute == 0:
        return DAY_MINUTES
    if not (0 <= hour < 24 and 0 <= minute < 60):
            return None
    return hour * 60 + minute


def parse_shift_cell(value: object) -> Tuple[List[Tuple[int, int]], Optional[str], str]:
    if value is None:
        return [], None, ""
    text = str(value).strip()
    if not text:
        return [], None, ""
    if text in REST_LABELS:
        return [], None, text
    parts = text.split()
    intervals: List[Tuple[int, int]] = []
    for part in parts:
        if "-" not in part:
            return [], f"无法识别排班格式: {text}", text
        start_token, end_token = part.split("-", 1)
        start = parse_time_to_minutes(start_token)
        end = parse_time_to_minutes(end_token)
        if start is None or end is None:
            return [], f"无效的时间范围: {part}", text
        if end < start:
            end += DAY_MINUTES
        intervals.append((start, end))
    return intervals, None, text


def clip_segment(start: int, end: int, clip_start: int, clip_end: Optional[int]) -> Optional[Tuple[int, int]]:
    upper = end if clip_end is None else min(end, clip_end)
    lower = max(start, clip_start)
    if upper <= lower:
        return None
    return lower, upper


def generate_segments(intervals: Iterable[Tuple[int, int]], rule: str) -> List[Tuple[int, int]]:
    segments: List[Tuple[int, int]] = []
    for start, end in intervals:
        if rule == "full":
            segments.append((start, end))
        elif rule == "until_midnight":
            clipped = clip_segment(start, end, 0, DAY_MINUTES)
            if clipped:
                segments.append(clipped)
        elif rule == "after_midnight":
            clipped = clip_segment(start, end, DAY_MINUTES, None)
            if clipped:
                segments.append(clipped)
        else:
            raise ValueError(f"未知的统计规则: {rule}")
    return segments


def format_rest_label(day_offset: int, label: str) -> str:
    if day_offset == 0:
        return label
    if day_offset == 1:
        return f"次日{label}"
    return f"第{day_offset}天{label}"


def rest_deductions_for_segment(start: int, end: int) -> Tuple[int, List[str]]:
    if end <= start:
        return 0, []
    total = 0
    labels: List[str] = []
    start_day = start // DAY_MINUTES
    end_day = (end - 1) // DAY_MINUTES
    for day_offset in range(start_day, end_day + 1):
        base = day_offset * DAY_MINUTES
        for rest_start, rest_end, label in REST_BREAKS:
            rs = base + rest_start
            re = base + rest_end
            overlap = min(end, re) - max(start, rs)
            if overlap > 0:
                total += overlap
                labels.append(format_rest_label(day_offset, label))
    return total, labels


def compute_overtime_detail(segments: Iterable[Tuple[int, int]]) -> Tuple[int, Tuple[str, ...]]:
    total_minutes = 0
    used_labels: List[str] = []
    for start, end in segments:
        if end <= start:
            continue
        duration = end - start
        deduction, labels = rest_deductions_for_segment(start, end)
        total_minutes += max(0, duration - deduction)
        used_labels.extend(labels)
    return total_minutes, tuple(used_labels)


def summarize_rest_labels(labels: Iterable[str]) -> str:
    ordered_counts: OrderedDict[str, int] = OrderedDict()
    for label in labels:
        ordered_counts[label] = ordered_counts.get(label, 0) + 1
    if not ordered_counts:
        return "无"
    parts: List[str] = []
    for label, count in ordered_counts.items():
        parts.append(f"{label}×{count}" if count > 1 else label)
    return "，".join(parts)


def build_result_comment(result: OvertimeResult) -> str:
    shift_text = result.shift_text or "-"
    summary = summarize_rest_labels(result.rest_labels)
    lines = [
        f"班次: {shift_text}",
        f"休息扣减: {summary}",
        f"加班工时: {result.hours} 小时",
    ]
    if result.error:
        lines.append(f"备注: {result.error}")
    return "\n".join(lines)


def calculate_overtime(value: object, rule: str) -> OvertimeResult:
    intervals, error, text = parse_shift_cell(value)
    if error:
        return OvertimeResult(0, (), text, error)
    if not intervals:
        return OvertimeResult(0, (), text)
    segments = generate_segments(intervals, rule)
    minutes, rest_labels = compute_overtime_detail(segments)
    hours = round(minutes / 60, 2)
    LOGGER.debug(
        "完成单元格计算: 班次=%s, 规则=%s, 时长=%.2f, 休息扣减=%s",
        text or "-",
        rule,
        hours,
        summarize_rest_labels(rest_labels),
    )
    return OvertimeResult(hours, rest_labels, text)


def extract_excel_date(value: object, *, date1904: bool) -> Optional[Tuple[int, int]]:
    if value is None:
        return None
    if isinstance(value, datetime):
        d = value.date()
        return d.month, d.day
    if isinstance(value, DateType):
        return value.month, value.day
    if isinstance(value, (int, float)):
        epoch = CALENDAR_MAC_1904 if date1904 else CALENDAR_WINDOWS_1900
        dt = from_excel(value, epoch=epoch)
        return dt.month, dt.day
    return None


def highlight_cell(cell, message: str) -> None:
    cell.fill = ERROR_FILL
    if cell.comment is None:
        cell.comment = Comment(text=message, author="Codex")
    LOGGER.warning("标记错误单元格 %s: %s", cell.coordinate, message)


def prompt_for_excel_file() -> Optional[Path]:
    if tk is None or filedialog is None:
        return None
    try:
        root = tk.Tk()
    except tk.TclError:
        return None
    root.withdraw()
    try:
        path = filedialog.askopenfilename(
            title="请选择考勤 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm"), ("所有文件", "*.*")],
        )
    finally:
        root.destroy()
    if not path:
        return None
    return Path(path)


def locate_default_input() -> Optional[Path]:
    current_dir = Path.cwd()
    candidates = [
        current_dir / "10月考勤.xlsx",
        current_dir / "files/10月考勤.xlsx",
        BASE_DIR / "10月考勤.xlsx",
        BASE_DIR / "files/10月考勤.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def setup_logging(base_dir: Path) -> Path:
    base_dir.mkdir(parents=True, exist_ok=True)
    log_path = base_dir / f"overtime_{datetime.now():%Y%m%d_%H%M%S}.log"

    if LOGGER.handlers:
        for handler in LOGGER.handlers[:]:
            LOGGER.removeHandler(handler)
            handler.close()

    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(formatter)
    LOGGER.addHandler(file_handler)

    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    LOGGER.addHandler(stream_handler)

    LOGGER.setLevel(logging.DEBUG)
    LOGGER.propagate = False

    LOGGER.info("日志文件: %s", log_path.resolve())
    return log_path


def main(input_path: Path, output_path: Path) -> None:
    if not input_path.exists():
        raise FileNotFoundError(f"未找到考勤文件: {input_path}")

    LOGGER.info("读取考勤文件: %s", input_path.resolve())
    workbook = load_workbook(input_path)
    sheet = workbook.active
    date1904 = workbook.epoch == CALENDAR_MAC_1904

    date_columns: dict[int, int] = {}
    for column_index, cell in enumerate(sheet[1], start=1):
        if column_index < 6:
            continue
        date_info = extract_excel_date(cell.value, date1904=date1904)
        if not date_info:
            continue
        month, day = date_info
        if month == 10 and day in TARGET_RULES:
            date_columns[day] = column_index
            LOGGER.info("识别到目标日期列: 10月%d日 -> 第 %d 列", day, column_index)

    missing = [day for day in TARGET_RULES if day not in date_columns]
    if missing:
        message = f"缺少目标日期列: {missing}"
        LOGGER.error(message)
        raise ValueError(message)

    max_row = sheet.max_row
    for day in [1, 2, 3, 5, 6]:
        schedule_col = date_columns[day]
        rule = TARGET_RULES[day]
        output_col = sheet.max_column + 1
        sheet.cell(row=1, column=output_col, value=f"{day}号加班")
        LOGGER.info("开始计算 %d 号加班列，源列索引 %d，新列索引 %d", day, schedule_col, output_col)

        for row in range(2, max_row + 1):
            source_cell = sheet.cell(row=row, column=schedule_col)
            result = calculate_overtime(source_cell.value, rule)
            output_cell = sheet.cell(row=row, column=output_col, value=result.hours)
            output_cell.comment = Comment(text=build_result_comment(result), author="Codex")
            if result.error:
                highlight_cell(source_cell, result.error)
                LOGGER.debug(
                    "排班解析错误 - 行 %d, 列 %d, 内容 %r, 原因: %s",
                    row,
                    schedule_col,
                    source_cell.value,
                    result.error,
                )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    LOGGER.info("结果已保存到: %s", output_path.resolve())
    print(f"加班统计已生成: {output_path.resolve()}")


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="读取考勤排班表，计算指定日期的加班工时并导出新表格。"
    )
    parser.add_argument(
        "-i",
        "--input",
        help="输入考勤表路径；若省略且支持图形界面，将弹窗选择文件。",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="输出文件路径（相对路径时写入 outputs/ 目录）",
    )
    parser.add_argument(
        "--no-dialog",
        action="store_true",
        help="禁用选择文件弹窗，未提供输入路径时回退到默认文件名。",
    )
    return parser.parse_args()


if __name__ == "__main__":
    cli_args = parse_arguments()
    log_path = setup_logging(OUTPUT_ROOT)
    LOGGER.info("程序启动")
    input_path: Optional[Path] = None
    if cli_args.input:
        input_path = Path(cli_args.input).expanduser()
    elif not cli_args.no_dialog:
        input_path = prompt_for_excel_file()
    if input_path is None:
        fallback = locate_default_input()
        if fallback is None:
            raise FileNotFoundError("未选择考勤文件，且未在默认位置找到 10月考勤.xlsx")
        input_path = fallback

    if cli_args.output:
        output_path = Path(cli_args.output).expanduser()
        if not output_path.is_absolute():
            output_path = OUTPUT_ROOT / output_path
    else:
        output_path = OUTPUT_ROOT / f"{input_path.stem}-加班统计.xlsx"
    LOGGER.info("输入文件: %s", input_path.resolve())
    LOGGER.info("输出文件: %s", output_path.resolve())
    try:
        main(input_path, output_path)
        LOGGER.info("程序结束")
        LOGGER.info("日志保存在: %s", log_path.resolve())
    except Exception as exc:  # pragma: no cover - top-level guard
        LOGGER.exception("程序异常终止: %s", exc)
        print(f"程序执行出错，详情请查看日志：{log_path.resolve()}", file=sys.stderr)
        sys.exit(1)
